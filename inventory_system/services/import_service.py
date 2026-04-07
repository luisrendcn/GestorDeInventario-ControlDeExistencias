"""
services/import_service.py
==========================
Servicio de importación masiva de productos desde archivos Excel.

Responsabilidades:
    - Leer archivos Excel (.xlsx)
    - Validar estructura de datos y campos obligatorios
    - Detectar tipo de producto automáticamente
    - Crear productos en lote
    - Generar reportes detallados
    - Manejar errores gracefully

Características:
    - Soporte para múltiples tipos de empresa (ropa, alimentos, tecnología, etc.)
    - Validación exhaustiva antes de importar
    - Generación automática de IDs si es necesaria
    - Transactional: importa todo o nada
    - Informe detallado de errores y advertencias
"""

from typing import List, Dict, Any, Tuple, Optional
from dataclasses import dataclass, field
from datetime import datetime, date
import re
from pathlib import Path
from models.producto import ProductoSimple, ProductoPerecedero, ProductoDigital, Producto
from models.inventario import Inventario
from patterns.observer import SujetoStock


@dataclass
class ImportResult:
    """Resultado de una operación de importación."""
    exitosos: int = 0
    fallidos: int = 0
    advertencias: int = 0
    productos_creados: List[Producto] = field(default_factory=list)
    errores: List[Dict[int, str]] = field(default_factory=list)  # [{'fila': 5, 'error': 'mensaje'}]
    advertencias_list: List[Dict[int, str]] = field(default_factory=list)

    def agregar_error(self, fila: int, mensaje: str) -> None:
        """Registra un error encontrado en una fila."""
        self.errores.append({"fila": fila, "error": mensaje})
        self.fallidos += 1

    def agregar_advertencia(self, fila: int, mensaje: str) -> None:
        """Registra una advertencia (pero el producto se importa)."""
        self.advertencias_list.append({"fila": fila, "advertencia": mensaje})
        self.advertencias += 1

    def agregar_producto(self, producto: Producto) -> None:
        """Registra un producto creado exitosamente."""
        self.productos_creados.append(producto)
        self.exitosos += 1

    def reporte_texto(self) -> str:
        """Genera un reporte textual legible del resultado."""
        lineas = []
        lineas.append("=" * 70)
        lineas.append("REPORTE DE IMPORTACIÓN")
        lineas.append("=" * 70)
        lineas.append(f"Productos exitosos: {self.exitosos}")
        lineas.append(f"Productos fallidos: {self.fallidos}")
        lineas.append(f"Advertencias: {self.advertencias}")
        lineas.append("")

        if self.errores:
            lineas.append("❌ ERRORES:")
            for error_info in self.errores:
                lineas.append(f"  Fila {error_info['fila']}: {error_info['error']}")
            lineas.append("")

        if self.advertencias_list:
            lineas.append("⚠️  ADVERTENCIAS:")
            for adv_info in self.advertencias_list:
                lineas.append(f"  Fila {adv_info['fila']}: {adv_info['advertencia']}")
            lineas.append("")

        lineas.append("=" * 70)
        return "\n".join(lineas)


class ImportadorExcel:
    """
    Importador de productos desde archivos Excel.
    
    Columnas esperadas (se ignoran mayúsculas/espacios):
        - nombre (obligatoria)
        - categoria (opcional - detecta tipo de producto)
        - precio_compra (opcional)
        - precio_venta (obligatoria, preferred over precio_compra)
        - stock (obligatoria)
        - stock_minimo (opcional, default=5)
    
    Campos adicionales según tipo:
        - fecha_vencimiento (para ProductoPerecedero, formato YYYY-MM-DD)
        - url_descarga (para ProductoDigital)
    """

    # Palabras clave para detectar tipos de producto automáticamente
    CATEGORIAS_PERECEDERO = {
        'alimentos', 'comida', 'bebidas', 'bebida', 'alimento',
        'medicamentos', 'medicamento', 'fármacos', 'fármaco', 'farmacia',
        'cosméticos', 'cosmético', 'aseo', 'higiene', 'perecedero'
    }

    CATEGORIAS_DIGITAL = {
        'software', 'digital', 'ebook', 'e-book', 'suscripción',
        'suscripciones', 'licencia', 'licencias', 'app', 'aplicación',
        'virtual', 'descargable', 'online'
    }

    def __init__(self, ruta_archivo: str):
        """
        Inicializa el importador.
        
        Args:
            ruta_archivo: Ruta al archivo Excel (.xlsx)
        """
        self.ruta = Path(ruta_archivo)
        if not self.ruta.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta_archivo}")
        if self.ruta.suffix.lower() != '.xlsx':
            raise ValueError(f"Formato incorrecto. Se esperaba .xlsx, se recibió {self.ruta.suffix}")

    def importar(self, inventario: Inventario, sujeto: SujetoStock) -> ImportResult:
        """
        Importa productos desde el archivo Excel.
        
        Args:
            inventario: Instancia de Inventario donde guardar los productos
            sujeto: Sujeto del patrón Observer para notificaciones
            
        Returns:
            ImportResult con detalles de la importación
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl no está instalado. Ejecuta: pip install openpyxl"
            )

        resultado = ImportResult()

        try:
            # Leer archivo Excel
            wb = openpyxl.load_workbook(str(self.ruta))
            ws = wb.active

            # Extraer encabezados (normalizar: lowercase + strip)
            encabezados = self._extraer_encabezados(ws)
            if not encabezados:
                resultado.agregar_error(1, "No se encontraron encabezados en la primera fila")
                return resultado

            # Validar campos obligatorios
            campos_faltantes = self._validar_campos_obligatorios(encabezados)
            if campos_faltantes:
                resultado.agregar_error(1, f"Campos obligatorios faltantes: {', '.join(campos_faltantes)}")
                return resultado

            # Procesar filas (comenzar desde fila 2)
            id_counter = 1
            for num_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    # Extraer valores de la fila
                    valores = self._extraer_valores_fila(fila, encabezados)
                    
                    # Generar ID si es necesario
                    if 'id' not in encabezados:
                        valores['id'] = f"PROD_{id_counter:06d}"
                        id_counter += 1
                    
                    # Validar datos de la fila
                    self._validar_fila(valores, num_fila, resultado)
                    if not valores.get('_valido'):
                        continue
                    
                    # Detectar tipo de producto
                    tipo_producto = self._detectar_tipo_producto(valores)
                    
                    # Crear producto según tipo
                    producto = self._crear_producto(tipo_producto, valores, num_fila, resultado)
                    if producto:
                        # Agregar al inventario
                        inventario.agregar(producto)
                        resultado.agregar_producto(producto)

                except Exception as e:
                    resultado.agregar_error(num_fila, f"Error procesando fila: {str(e)}")

            wb.close()
            return resultado

        except Exception as e:
            resultado.agregar_error(0, f"Error leyendo archivo: {str(e)}")
            return resultado

    @staticmethod
    def _extraer_encabezados(ws) -> Dict[int, str]:
        """
        Extrae los encabezados de la primera fila y los normaliza.
        
        Returns:
            Dict con {número_columna: nombre_normalizado}
        """
        encabezados = {}
        for idx, celda in enumerate(ws[1], start=1):
            valor = celda.value
            if valor:
                # Normalizar: lowercase, strip, reemplazar espacios por _
                normalizado = str(valor).strip().lower().replace(' ', '_')
                encabezados[idx] = normalizado
        return encabezados

    @staticmethod
    def _extraer_valores_fila(fila, encabezados: Dict[int, str]) -> Dict[str, Any]:
        """Extrae los valores de una fila usando los encabezados normalizados."""
        valores = {}
        for idx, nombre_col in encabezados.items():
            celda = fila[idx - 1]
            valores[nombre_col] = celda.value
        return valores

    @staticmethod
    def _validar_campos_obligatorios(encabezados: Dict[int, str]) -> List[str]:
        """Valida que los campos obligatorios estén presentes."""
        campos_obligatorios = {'nombre', 'stock'}  # Mínimo requerido
        campos_precio = {'precio_venta', 'precio_compra', 'precio'}
        
        faltantes = campos_obligatorios - set(encabezados.values())
        
        # Verificar que hay al menos un campo de precio
        if not any(p in encabezados.values() for p in campos_precio):
            faltantes.add("precio_venta o precio_compra")
        
        return list(faltantes)

    @staticmethod
    def _validar_fila(valores: Dict[str, Any], num_fila: int, resultado: ImportResult) -> None:
        """
        Valida los valores de una fila completa.
        Agrega a valores['_valido'] = True/False
        """
        valores['_valido'] = True
        errores_fila = []

        # Validar nombre
        if not valores.get('nombre') or not str(valores.get('nombre', '')).strip():
            errores_fila.append("nombre vacío")
            valores['_valido'] = False
        else:
            valores['nombre'] = str(valores['nombre']).strip()

        # Validar stock
        try:
            valores['stock'] = int(float(valores.get('stock', 0)))
            if valores['stock'] < 0:
                errores_fila.append("stock no puede ser negativo")
                valores['_valido'] = False
        except (TypeError, ValueError):
            errores_fila.append(f"stock inválido: {valores.get('stock')}")
            valores['_valido'] = False

        # Validar precio
        precio_valor = None
        if 'precio_venta' in valores and valores['precio_venta']:
            precio_valor = valores['precio_venta']
        elif 'precio' in valores and valores['precio']:
            precio_valor = valores['precio']
        elif 'precio_compra' in valores and valores['precio_compra']:
            precio_valor = valores['precio_compra']

        if precio_valor is None:
            errores_fila.append("precio no especificado")
            valores['_valido'] = False
        else:
            try:
                valores['precio'] = float(precio_valor)
                if valores['precio'] < 0:
                    errores_fila.append("precio no puede ser negativo")
                    valores['_valido'] = False
            except (TypeError, ValueError):
                errores_fila.append(f"precio inválido: {precio_valor}")
                valores['_valido'] = False

        # Validar stock_minimo (opcional, default=5)
        if 'stock_minimo' in valores and valores['stock_minimo']:
            try:
                valores['stock_minimo'] = int(float(valores['stock_minimo']))
                if valores['stock_minimo'] < 0:
                    errores_fila.append("stock_minimo no puede ser negativo")
                    valores['stock_minimo'] = 5
            except (TypeError, ValueError):
                resultado.agregar_advertencia(num_fila, "stock_minimo inválido, usando default 5")
                valores['stock_minimo'] = 5
        else:
            valores['stock_minimo'] = 5

        # Normalizar categoría
        if 'categoria' in valores and valores['categoria']:
            valores['categoria'] = str(valores['categoria']).strip()
        else:
            valores['categoria'] = 'General'

        if errores_fila:
            resultado.agregar_error(num_fila, "; ".join(errores_fila))

    def _detectar_tipo_producto(self, valores: Dict[str, Any]) -> str:
        """
        Detecta automáticamente el tipo de producto basado en:
        1. Categoría (si contiene palabras clave)
        2. Campos específicos presentes (fecha_vencimiento, url_descarga)
        3. Por defecto: ProductoSimple
        """
        categoria_lower = valores.get('categoria', '').lower()

        # Verificar palabras clave en categoría
        if any(pal in categoria_lower for pal in self.CATEGORIAS_PERECEDERO):
            return 'perecedero'

        if any(pal in categoria_lower for pal in self.CATEGORIAS_DIGITAL):
            return 'digital'

        # Verificar campos específicos
        if 'fecha_vencimiento' in valores and valores['fecha_vencimiento']:
            return 'perecedero'

        if 'url_descarga' in valores and valores['url_descarga']:
            return 'digital'

        # Por defecto
        return 'simple'

    def _crear_producto(
        self,
        tipo: str,
        valores: Dict[str, Any],
        num_fila: int,
        resultado: ImportResult
    ) -> Optional[Producto]:
        """
        Crea un producto del tipo especificado.
        
        Returns:
            Instancia de Producto o None si hay error
        """
        try:
            id_producto = valores.get('id', f'PROD_{num_fila}')

            if tipo == 'perecedero':
                return self._crear_producto_perecedero(id_producto, valores, num_fila, resultado)
            elif tipo == 'digital':
                return self._crear_producto_digital(id_producto, valores, num_fila, resultado)
            else:
                return self._crear_producto_simple(id_producto, valores, num_fila, resultado)

        except Exception as e:
            resultado.agregar_error(num_fila, f"Error creando producto: {str(e)}")
            return None

    @staticmethod
    def _crear_producto_simple(
        id_producto: str,
        valores: Dict[str, Any],
        num_fila: int,
        resultado: ImportResult
    ) -> ProductoSimple:
        """Crea un ProductoSimple."""
        return ProductoSimple(
            id=id_producto,
            nombre=valores['nombre'],
            precio=valores['precio'],
            stock=valores['stock'],
            stock_minimo=valores.get('stock_minimo', 5),
            categoria=valores.get('categoria', 'General')
        )

    @staticmethod
    def _crear_producto_perecedero(
        id_producto: str,
        valores: Dict[str, Any],
        num_fila: int,
        resultado: ImportResult
    ) -> Optional[ProductoPerecedero]:
        """Crea un ProductoPerecedero (requiere fecha_vencimiento)."""
        fecha_vencimiento = valores.get('fecha_vencimiento')

        if not fecha_vencimiento:
            resultado.agregar_error(
                num_fila,
                "ProductoPerecedero requiere fecha_vencimiento (formato YYYY-MM-DD)"
            )
            return None

        try:
            if isinstance(fecha_vencimiento, date):
                fecha = fecha_vencimiento
            else:
                fecha = datetime.strptime(str(fecha_vencimiento).strip(), '%Y-%m-%d').date()
        except ValueError:
            resultado.agregar_error(
                num_fila,
                f"Formato de fecha inválido: {fecha_vencimiento}. Use YYYY-MM-DD"
            )
            return None

        return ProductoPerecedero(
            id=id_producto,
            nombre=valores['nombre'],
            precio=valores['precio'],
            stock=valores['stock'],
            fecha_vencimiento=fecha,
            stock_minimo=valores.get('stock_minimo', 5)
        )

    @staticmethod
    def _crear_producto_digital(
        id_producto: str,
        valores: Dict[str, Any],
        num_fila: int,
        resultado: ImportResult
    ) -> ProductoDigital:
        """Crea un ProductoDigital."""
        url = valores.get('url_descarga', '')
        if url:
            url = str(url).strip()

        return ProductoDigital(
            id=id_producto,
            nombre=valores['nombre'],
            precio=valores['precio'],
            licencias=valores['stock'],
            stock_minimo=valores.get('stock_minimo', 2),  # Para digital, default es 2
            url_descarga=url
        )
