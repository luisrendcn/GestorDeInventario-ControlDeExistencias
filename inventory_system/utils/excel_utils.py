"""
utils/excel_utils.py
===================
Utilidades para trabajar con archivos Excel en importación/exportación.

Funcionalidades:
    - Generar plantillas de ejemplo
    - Crear archivos de prueba
    - Validar estructura de Excel
"""

from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional, List


def generar_plantilla_excel(ruta_salida: str, tipo_empresa: str = 'general') -> None:
    """
    Genera un archivo Excel de plantilla con ejemplos.
    
    Args:
        ruta_salida: Ruta donde guardar el archivo (.xlsx)
        tipo_empresa: Tipo de empresa ('general', 'alimentos', 'ropa', 'tecnologia')
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Definir estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezados
    encabezados = ["nombre", "categoria", "precio_compra", "precio_venta", "stock", "stock_minimo"]
    
    # Agregar campos específicos según tipo
    if tipo_empresa == 'alimentos':
        encabezados.append("fecha_vencimiento")
    elif tipo_empresa == 'tecnologia':
        encabezados.append("url_descarga")

    # Escribir encabezados
    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col)
        celda.value = encabezado
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = center_alignment
        celda.border = border

    # Agregar filas de ejemplo según tipo
    ejemplos = _generar_ejemplos(tipo_empresa)
    
    for fila_idx, ejemplo in enumerate(ejemplos, start=2):
        for col_idx, valor in enumerate(ejemplo, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx)
            celda.value = valor
            celda.alignment = Alignment(horizontal="left", vertical="center")
            celda.border = border

    # Ajustar ancho de columnas
    anchos = [20, 15, 12, 12, 8, 12]
    if tipo_empresa in ['alimentos', 'tecnologia']:
        anchos.append(25)

    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + col)].width = ancho

    # Agregar instrucciones en otra hoja
    ws_info = wb.create_sheet("Instrucciones")
    instrucciones = _generar_instrucciones(tipo_empresa)
    for idx, linea in enumerate(instrucciones, start=1):
        celda = ws_info.cell(row=idx, column=1)
        celda.value = linea
        if idx == 1:
            celda.font = Font(bold=True, size=12)

    ws_info.column_dimensions['A'].width = 100

    # Guardar
    ruta = Path(ruta_salida)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta))
    print(f"✅ Plantilla creada: {ruta}")


def _generar_ejemplos(tipo_empresa: str) -> List[List]:
    """Genera filas de ejemplo según el tipo de empresa."""
    hoy = datetime.now().date()
    
    if tipo_empresa == 'alimentos':
        # Ejemplo para empresa de alimentos
        fecha_vto1 = (hoy + timedelta(days=30)).strftime('%Y-%m-%d')
        fecha_vto2 = (hoy + timedelta(days=60)).strftime('%Y-%m-%d')
        fecha_vto3 = (hoy + timedelta(days=90)).strftime('%Y-%m-%d')
        
        return [
            ["Arroz Integral 1kg", "Alimentos", 1.50, 3.99, 100, 20, fecha_vto1],
            ["Leche Entera 1L", "Bebidas", 0.80, 1.99, 50, 15, fecha_vto2],
            ["Pollo Premium kg", "Carnes", 4.50, 8.99, 30, 10, fecha_vto3],
            ["Pan Integral", "Panadería", 0.50, 1.49, 80, 25, fecha_vto1],
            ["Queso Fresco 500g", "Lácteos", 3.00, 7.49, 25, 8, fecha_vto2],
        ]
    
    elif tipo_empresa == 'ropa':
        return [
            ["Camiseta Básica", "Camisetas", 2.00, 12.99, 150, 30, ""],
            ["Pantalón Jeans", "Pantalones", 8.00, 34.99, 50, 10, ""],
            ["Sudadera Deportiva", "Sudaderas", 5.00, 24.99, 40, 8, ""],
            ["Conjunto Deportivo", "Conjuntos", 10.00, 39.99, 30, 5, ""],
            ["Calcetines (paquete 6)", "Accesorios", 0.80, 4.99, 100, 20, ""],
        ]
    
    elif tipo_empresa == 'tecnologia':
        url1 = "https://example.com/software1"
        url2 = "https://example.com/software2"
        url3 = "https://example.com/ebook-python"
        
        return [
            ["Licencia Software X", "Software", 50.00, 149.99, 10, 2, url1],
            ["Suscripción Cloud Annual", "Suscripciones", 100.00, 299.99, 5, 1, url2],
            ["E-book Python Avanzado", "E-books", 2.00, 14.99, 999, 10, url3],
            ["Plugin VSCode Premium", "Extensiones", 5.00, 19.99, 50, 5, url1],
            ["Template Diseño Web", "Plantillas", 10.00, 29.99, 100, 20, url2],
        ]
    
    else:  # general
        return [
            ["Producto Ejemplo 1", "General", 5.00, 15.99, 100, 20, ""],
            ["Producto Ejemplo 2", "Electrónica", 15.00, 49.99, 50, 10, ""],
            ["Producto Ejemplo 3", "General", 2.50, 9.99, 200, 40, ""],
            ["Producto Ejemplo 4", "Accesorios", 3.00, 12.99, 75, 15, ""],
            ["Producto Ejemplo 5", "General", 10.00, 29.99, 30, 5, ""],
        ]


def _generar_instrucciones(tipo_empresa: str) -> List[str]:
    """Genera instrucciones según el tipo de empresa."""
    instrucciones = [
        "INSTRUCCIONES PARA IMPORTAR PRODUCTOS",
        "",
        "1. COLUMNAS OBLIGATORIAS:",
        "   - nombre: Nombre del producto",
        "   - precio_venta: Precio al que se vende",
        "   - stock: Cantidad disponible",
        "",
        "2. COLUMNAS OPCIONALES:",
        "   - categoria: Clasificación del producto (General, Alimentos, etc.)",
        "   - precio_compra: Precio de costo",
        "   - stock_minimo: Cantidad mínima antes de alertar (default: 5)",
    ]
    
    if tipo_empresa == 'alimentos':
        instrucciones.extend([
            "",
            "3. PARA ALIMENTOS (Productos Perecederos):",
            "   - fecha_vencimiento: OBLIGATORIA en formato YYYY-MM-DD",
            "   - El sistema detecta automáticamente que es perecedero por:",
            "     * La categoría contiene palabras clave (alimentos, bebidas, etc.)",
            "     * O si especificas fecha_vencimiento",
        ])
    
    elif tipo_empresa == 'tecnologia':
        instrucciones.extend([
            "",
            "3. PARA TECNOLOGÍA (Productos Digitales):",
            "   - url_descarga: URL donde descargar el producto",
            "   - El sistema detecta automáticamente que es digital por:",
            "     * La categoría contiene palabras clave (software, digital, etc.)",
            "     * O si especificas url_descarga",
        ])
    
    instrucciones.extend([
        "",
        "4. FORMATO DE DATOS:",
        "   - Precios: números con punto decimal (ej: 19.99)",
        "   - Stock: números enteros (ej: 100)",
        "   - Fechas: YYYY-MM-DD (ej: 2025-12-31)",
        "",
        "5. VALIDACIONES:",
        "   - Todos los precios deben ser positivos",
        "   - Stock no pueden ser negativos",
        "   - Nombres no pueden estar vacíos",
        "",
        "6. DESPUÉS DE IMPORTAR:",
        "   - Recibirás un reporte con:",
        "     * Total de productos importados exitosamente",
        "     * Errores encontrados (con número de fila)",
        "     * Advertencias (pero se importan de todas formas)",
    ])
    
    return instrucciones


def validar_estructura_excel(ruta_archivo: str, campos_obligatorios: Optional[List[str]] = None) -> bool:
    """
    Valida que un archivo Excel tenga la estructura correcta.
    
    Args:
        ruta_archivo: Ruta al archivo Excel
        campos_obligatorios: Campos que deben estar presentes
        
    Returns:
        True si la estructura es válida
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl no está instalado")

    ruta = Path(ruta_archivo)
    if not ruta.exists():
        print(f"❌ Archivo no encontrado: {ruta_archivo}")
        return False

    if ruta.suffix.lower() != '.xlsx':
        print(f"❌ Formato incorrecto. Se esperaba .xlsx")
        return False

    try:
        wb = openpyxl.load_workbook(str(ruta))
        ws = wb.active

        # Extraer encabezados
        encabezados = []
        for celda in ws[1]:
            if celda.value:
                encabezados.append(str(celda.value).strip().lower())

        if not encabezados:
            print("❌ No se encontraron encabezados")
            return False

        # Validar campos obligatorios
        if campos_obligatorios:
            faltantes = set(campos_obligatorios) - set(encabezados)
            if faltantes:
                print(f"❌ Campos obligatorios faltantes: {', '.join(faltantes)}")
                return False

        # Contar filas de datos
        num_filas = ws.max_row - 1  # Excluir encabezado
        print(f"✅ Estructura válida. Encabezados: {len(encabezados)} | Filas: {num_filas}")
        print(f"   Columnas: {', '.join(encabezados)}")

        wb.close()
        return True

    except Exception as e:
        print(f"❌ Error validando archivo: {str(e)}")
        return False
